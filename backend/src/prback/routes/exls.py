'''
This module contains routes for generating Excel reports of ACL requests. It provides endpoints to create color-coded Excel files 
based on the status and system type of ACL requests, allowing users to easily visualize and analyze their submissions.
The generated Excel files include conditional formatting to highlight different statuses (e.g., Pending, Approved, Rejected) a
nd system types (e.g., "Others" for manual entries).
'''
import logging
from flask import Blueprint
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from flask import jsonify, request, send_file
from ..models import ACLRequest
from ..guards.roleguard import token_required

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

exls_bp = Blueprint('exls', __name__)


@exls_bp.route('/api/v1/generate-xlsx/submission', methods=['POST'])
@token_required()
def generate_submission_xlsx(current_user):
    '''Generate color-coded Excel file for a specific submission'''
    try:
        data = request.get_json()

        # Expect array of request IDs or request data from current submission
        if 'request_ids' in data:
            # Fetch by IDs
            request_ids = data['request_ids']
            requests = ACLRequest.query.filter(
                ACLRequest.id.in_(request_ids)).all()
        elif 'requests' in data:
            # Create from submitted data (for immediate download after submission)
            requests = data['requests']
        else:
            return jsonify({'error': 'No requests provided'}), 400

        if not requests:
            return jsonify({'error': 'No ACL requests found'}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ACL Requests"

        # Define colors for different statuses and system types
        status_colors = {
            'Pending': 'FFF3CD',      # Yellow
            'Approved': 'D4EDDA',     # Green
            'Rejected': 'F8D7DA',     # Red
            'Completed': 'D1ECF1'     # Blue
        }

        system_type_colors = {
            'Others': 'FFE699',      # Light orange for manual "Others"
            'Template': 'E2EFDA',    # Light green for template-based entries
        }

        # Color for "Others" system type
        others_row_color = 'FFE699'  # Light orange for "Others"

        # Define header style
        header_fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Request ID', 'Requester', 'System Type', 'Category', 'Source IP',
            'Source Host', 'Destination IP', 'Destination Host',
            'Service', 'Reason', 'Status', 'Created At'
        ]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows with conditional formatting
        for row_idx, req in enumerate(requests, start=2):
            # Handle both ACLRequest objects and dict data
            if isinstance(req, dict):
                request_data = [
                    req.get('id', 'N/A'),
                    req.get('requester', current_user.username),
                    req.get('system_type', ''),
                    req.get('category', ''),
                    req.get('sourceIP', ''),
                    req.get('sourceHost', ''),
                    req.get('destinationIP', ''),
                    req.get('destinationHost', ''),
                    req.get('service', ''),
                    req.get('description', ''),
                    req.get('status', 'Pending'),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
                system_type = req.get('system_type', '')
                status = req.get('status', 'Pending')
            else:
                request_data = [
                    req.id,
                    req.requester,
                    req.system_type,
                    req.category,
                    req.source_ip,
                    req.source_host,
                    req.destination_ip,
                    req.destination_host,
                    req.service,
                    req.reason,
                    req.status,
                    req.created_at.strftime(
                        "%Y-%m-%d %H:%M:%S") if req.created_at else 'N/A'
                ]
                system_type = req.system_type
                status = req.status

            ws.append(request_data)

            # NEW: Determine row background color based on system_type
            is_others_system = system_type == 'Others'
            is_template_system = system_type == 'Template'

            # Get the status color for this row
            status_color = status_colors.get(status, 'FFFFFF')
            status_fill = PatternFill(
                start_color=status_color, end_color=status_color, fill_type='solid')

            # NEW: Row fill for "Others" system type
            others_fill = PatternFill(
                start_color=others_row_color, end_color=others_row_color, fill_type='solid')

            if is_others_system:
                row_fill = PatternFill(
                    start_color=system_type_colors['Others'],
                    end_color=system_type_colors['Others'],
                    fill_type='solid'
                )
            elif is_template_system:
                row_fill = PatternFill(
                    start_color=system_type_colors['Template'],
                    end_color=system_type_colors['Template'],
                    fill_type='solid'
                )
            else:
                row_fill = None

            # Apply styling to each cell in the row
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                if row_fill and col_idx != 11:
                    cell.fill = row_fill
                    if is_others_system:
                        cell.font = Font(italic=True)
                    elif is_template_system:
                        cell.font = Font(bold=True, color='2F5233')

                # If this is an "Others" row, apply yellow/orange background to entire row
                if is_others_system and col_idx != 11:  # All columns except Status
                    cell.fill = others_fill
                    # Add italic font to distinguish "Others" entries
                    cell.font = Font(italic=True)

                # Color code the Status column (column 11)
                if col_idx == 11:
                    cell.fill = status_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

                # Highlight IP addresses (columns 5, 7) - only if NOT "Others"
                elif col_idx in [5, 7] and not is_others_system:
                    cell.fill = PatternFill(
                        start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.font = Font(name='Courier New')

                # Highlight Service (column 9) - only if NOT "Others"
                elif col_idx == 9 and not is_others_system:
                    cell.fill = PatternFill(
                        start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    cell.font = Font(name='Courier New')

                # NEW: Highlight System Type column (column 3) for "Others"
                elif col_idx == 3 and is_others_system:
                    cell.fill = PatternFill(
                        start_color='FF9900', end_color='FF9900', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = 'A2'

        # Add filters to header row
        ws.auto_filter.ref = ws.dimensions

        # Add legend for color coding
        legend_row = len(requests) + 3
        ws[f'A{legend_row}'] = 'Color Legend:'
        ws[f'A{legend_row}'].font = Font(bold=True)

        ws[f'A{legend_row + 1}'] = 'Green highlight = "Template" system type (from saved template)'
        ws[f'A{legend_row + 1}'].fill = PatternFill(
            start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'
        )

        ws[f'A{legend_row + 2}'] = 'Orange highlight = "Others" system type (manual entry)'
        ws[f'A{legend_row + 2}'].fill = PatternFill(
            start_color='FFE699', end_color='FFE699', fill_type='solid'
        )

        ws[f'A{legend_row + 2}'] = 'Yellow status = Pending'
        ws[f'A{legend_row + 2}'].fill = PatternFill(
            start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

        ws[f'A{legend_row + 3}'] = 'Green status = Approved'
        ws[f'A{legend_row + 3}'].fill = PatternFill(
            start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Filename includes submission timestamp and request count
        submission_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'acl_submission_{len(requests)}requests_{submission_time}.xlsx'

        return send_file(
            file_stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error("Error generating submission Excel", exc_info=True)
        return jsonify({'error': 'Failed to generate Excel report'}), 500


@exls_bp.route('/api/v1/generate-xlsx', methods=['GET'])
@token_required()
def generate_xlsx_enhanced(current_user):
    '''Generate color-coded Excel file after submission'''
    try:
        if current_user.role != 'admin':
            requests = ACLRequest.query.filter_by(requester=current_user.username).order_by(
                ACLRequest.created_at.desc()).all()
        else:
            requests = ACLRequest.query.order_by(
                ACLRequest.created_at.desc()).all()

        if not requests:
            return jsonify({'error': 'No ACL requests found to generate report'}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All ACL Requests"

        # Define colors for different statuses
        status_colors = {
            'Pending': 'FFF3CD',      # Yellow
            'Approved': 'D4EDDA',     # Green
            'Rejected': 'F8D7DA',     # Red
            'Completed': 'D1ECF1'     # Blue
        }
        others_row_color = 'FFE699'

        # Define header style
        header_fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'Request ID', 'Requester', 'System Type', 'Category', 'Source IP',
            'Source Host', 'Destination IP', 'Destination Host',
            'Service', 'Reason', 'Status', 'Created At'
        ]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows with conditional formatting
        for row_idx, req in enumerate(requests, start=2):
            ws.append([
                req.id,
                req.requester,
                req.system_type,
                req.category,
                req.source_ip,
                req.source_host,
                req.destination_ip,
                req.destination_host,
                req.service,
                req.reason,
                req.status,
                req.created_at.strftime(
                    "%Y-%m-%d %H:%M:%S") if req.created_at else 'N/A'
            ])

            # Get the status color for 'others' row
            is_others_system = system_type == 'Others'

            # Get the status color for this row
            status_color = status_colors.get(status, 'FFFFFF')
            status_fill = PatternFill(
                start_color=status_color, end_color=status_color, fill_type='solid')

            # Row fill for "Others" system type
            others_fill = PatternFill(
                start_color=others_row_color, end_color=others_row_color, fill_type='solid')

            # Apply styling to each cell in the row
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Color code the Status column (column 11)
                if col_idx == 11:
                    cell.fill = status_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

                # Highlight IP addresses (columns 5, 7)
                elif col_idx in [5, 7]:
                    cell.fill = PatternFill(
                        start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.font = Font(name='Courier New')

                # Highlight Service (column 9)
                elif col_idx == 9:
                    cell.fill = PatternFill(
                        start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    cell.font = Font(name='Courier New')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = 'A2'

        # Add filters to header row
        ws.auto_filter.ref = ws.dimensions

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename = f'acl_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            file_stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error("Error generating Excel", exc_info=True)
        return jsonify({'error': 'Failed to generate Excel report'}), 500